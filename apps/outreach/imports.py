import csv
import io
from pathlib import Path

from django.conf import settings
from rest_framework.exceptions import ValidationError

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


CONSENT_FIELD_ALIASES = {
    "client_id": {"client_id", "client", "id", "crm_id"},
    "phone": {"phone", "телефон", "номер", "whatsapp", "mobile"},
    "email": {"email", "почта", "e-mail"},
    "channel": {"channel", "канал", "source_channel"},
    "status": {"status", "статус", "consent", "согласие"},
    "source": {"source", "источник", "evidence_source"},
    "note": {"note", "notes", "комментарий", "заметка", "основание"},
}


def read_consent_upload(upload):
    extension = Path(getattr(upload, "name", "")).suffix.lower()
    if extension == ".csv":
        return _read_csv(upload)
    if extension == ".xlsx":
        return _read_xlsx(upload)
    raise ValidationError({"file": "Only CSV and XLSX files are supported."})


def normalize_consent_rows(rows):
    normalized = []
    for row in rows:
        clean = {str(key or "").strip(): _cell_to_string(value) for key, value in (row or {}).items() if str(key or "").strip()}
        if not any(clean.values()):
            continue
        normalized.append(_normalize_consent_row(clean))
    if not normalized:
        raise ValidationError({"file": "Import file has headers but no data rows."})
    if len(normalized) > settings.IMPORT_MAX_ROWS:
        raise ValidationError({"file": f"Import file has too many rows. Maximum is {settings.IMPORT_MAX_ROWS}."})
    return normalized


def _read_csv(upload):
    upload.seek(0)
    text = upload.read()
    if isinstance(text, bytes):
        text = text.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise ValidationError({"file": "Import file is empty."})
    return normalize_consent_rows(list(reader))


def _read_xlsx(upload):
    if load_workbook is None:
        raise ValidationError({"file": "XLSX import requires openpyxl."})
    upload.seek(0)
    workbook = load_workbook(upload, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    if not rows:
        raise ValidationError({"file": "Import file is empty."})
    headers = [_cell_to_string(value) for value in rows[0]]
    return normalize_consent_rows([{headers[index]: value for index, value in enumerate(values) if index < len(headers)} for values in rows[1:]])


def _normalize_consent_row(row):
    normalized_headers = {_normalize_header(header): header for header in row}
    result = {}
    for field, aliases in CONSENT_FIELD_ALIASES.items():
        for alias in aliases:
            header = normalized_headers.get(_normalize_header(alias))
            if header:
                result[field] = row.get(header, "")
                break
    return result


def _normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def _cell_to_string(value):
    if value is None:
        return ""
    return str(value).strip()
