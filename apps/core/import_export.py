import csv
import hashlib
import io
import json
from decimal import Decimal, InvalidOperation
from pathlib import Path

from django.conf import settings
from django.db import transaction
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from rest_framework.exceptions import ValidationError

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - local env may install requirements after code checkout.
    load_workbook = None

from apps.clients.models import Client
from apps.clients.services import duplicate_payload, find_duplicate_clients
from apps.core.audit import write_audit_log
from apps.core.csv_safety import safe_csv_cell
from apps.core.models import AuditLog, ImportJob
from apps.crm.models import Deal
from apps.integrations.connectors import normalize_business_event
from apps.integrations.models import BusinessConnector, BusinessEvent, ConnectorSyncRun
from apps.leads.models import Lead
from apps.services.models import Service


CLIENT_FIELD_ALIASES = {
    "full_name": ["full_name", "name", "имя", "фио", "клиент", "client", "contact"],
    "phone": ["phone", "телефон", "номер", "mobile", "whatsapp"],
    "email": ["email", "почта", "e-mail"],
    "source": ["source", "источник"],
    "notes": ["notes", "note", "заметки", "комментарий"],
}

CLIENT_FIELDS = ["full_name", "phone", "email", "source", "notes"]

LEAD_FIELD_ALIASES = {
    "full_name": ["full_name", "name", "имя", "фио", "клиент", "client", "contact"],
    "phone": ["phone", "телефон", "номер", "mobile", "whatsapp"],
    "email": ["email", "почта", "e-mail"],
    "service_name": ["service_name", "service", "услуга", "товар", "интерес"],
    "source": ["source", "источник", "channel", "канал"],
    "message": ["message", "сообщение", "запрос", "comment", "комментарий"],
    "status": ["status", "статус"],
}

SALES_FIELD_ALIASES = {
    "external_id": ["external_id", "order_id", "sale_id", "номер_заказа", "номер", "id"],
    "occurred_at": ["occurred_at", "date", "datetime", "дата", "дата_продажи"],
    "client_name": ["client_name", "client", "клиент", "покупатель", "фио"],
    "phone": ["phone", "телефон", "номер", "mobile", "whatsapp"],
    "item_name": ["item_name", "service", "product", "товар", "услуга", "позиция", "name"],
    "quantity": ["quantity", "qty", "количество", "шт"],
    "amount": ["amount", "total", "price", "sum", "выручка", "сумма", "стоимость"],
    "source": ["source", "источник", "channel", "канал"],
    "notes": ["notes", "note", "комментарий", "заметки"],
}

CATALOG_FIELD_ALIASES = {
    "item_type": ["item_type", "type", "тип", "категория_типа"],
    "sku": ["sku", "артикул", "code", "код"],
    "name": ["name", "название", "товар", "услуга", "позиция"],
    "description": ["description", "описание", "notes", "заметки"],
    "duration_minutes": ["duration_minutes", "duration", "длительность", "минуты"],
    "price_from": ["price_from", "price", "цена", "стоимость"],
    "stock_quantity": ["stock_quantity", "stock", "остаток", "quantity", "qty"],
    "source": ["source", "источник"],
}

FIELD_ALIASES_BY_ENTITY = {
    ImportJob.EntityTypes.CLIENTS: CLIENT_FIELD_ALIASES,
    ImportJob.EntityTypes.LEADS: LEAD_FIELD_ALIASES,
    ImportJob.EntityTypes.SALES: SALES_FIELD_ALIASES,
    ImportJob.EntityTypes.CATALOG: CATALOG_FIELD_ALIASES,
}

FIELDS_BY_ENTITY = {
    ImportJob.EntityTypes.CLIENTS: CLIENT_FIELDS,
    ImportJob.EntityTypes.LEADS: list(LEAD_FIELD_ALIASES.keys()),
    ImportJob.EntityTypes.SALES: list(SALES_FIELD_ALIASES.keys()),
    ImportJob.EntityTypes.CATALOG: list(CATALOG_FIELD_ALIASES.keys()),
}

IMPORT_TEMPLATES = {
    ImportJob.EntityTypes.CLIENTS: {
        "headers": CLIENT_FIELDS,
        "row": ["Алия Иванова", "+77015550101", "aliya@example.com", "manual", "VIP клиент"],
    },
    ImportJob.EntityTypes.LEADS: {
        "headers": list(LEAD_FIELD_ALIASES.keys()),
        "row": ["Алия Иванова", "+77015550101", "aliya@example.com", "Консультация", "landing", "Хочу записаться", "new"],
    },
    ImportJob.EntityTypes.SALES: {
        "headers": list(SALES_FIELD_ALIASES.keys()),
        "row": ["sale-001", "2026-05-22T10:00:00+05:00", "Алия Иванова", "+77015550101", "Консультация", "1", "15000", "manual", "Оплачено"],
    },
    ImportJob.EntityTypes.CATALOG: {
        "headers": list(CATALOG_FIELD_ALIASES.keys()),
        "row": ["service", "CONSULT-30", "Консультация", "Первичная консультация", "30", "15000", "", "manual"],
    },
}


def build_import_preview(job: ImportJob, mapping=None):
    rows = read_tabular_file(job.source_file.path)
    headers = list(rows[0].keys()) if rows else []
    mapping = mapping or guess_mapping(headers, aliases_for_entity(job.entity_type))
    preview_rows = rows[: settings.IMPORT_PREVIEW_ROWS]
    duplicates = duplicate_preview(job, preview_rows, mapping) if job.entity_type == ImportJob.EntityTypes.CLIENTS else []
    row_errors = validate_import_rows(job.entity_type, rows, mapping)
    job.mapping_json = mapping
    job.preview_json = {
        "headers": headers,
        "rows": preview_rows,
        "import_summary": {
            "total_rows": len(rows),
            "preview_rows": len(preview_rows),
            "errors": len(row_errors),
            "duplicates": len(duplicates),
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "ready": not row_errors,
        },
    }
    job.duplicates_json = {"rows": duplicates}
    job.errors_json = {"rows": row_errors}
    job.total_rows = len(rows)
    job.status = ImportJob.Statuses.PREVIEWED
    job.error = readable_import_error(row_errors)
    job.save(update_fields=["mapping_json", "preview_json", "duplicates_json", "errors_json", "total_rows", "status", "error", "updated_at"])
    return job


@transaction.atomic
def confirm_import(job: ImportJob, request):
    mapping = job.mapping_json or {}
    if not mapping:
        raise ValidationError("Run preview before confirming import.")
    if (job.errors_json or {}).get("rows"):
        raise ValidationError("Fix import errors before confirming.")
    rows = read_tabular_file(job.source_file.path)
    connector = ensure_excel_csv_connector(job)
    sync_run = ConnectorSyncRun.objects.create(
        business=job.business,
        connector=connector,
        mode=ConnectorSyncRun.Modes.MANUAL,
        status=ConnectorSyncRun.Statuses.RUNNING,
        started_at=timezone.now(),
    )
    summary = {"created": 0, "updated": 0, "skipped": 0}
    try:
        for row in rows:
            payload = mapped_payload(row, mapping, job.entity_type)
            if job.entity_type == ImportJob.EntityTypes.CLIENTS:
                result = import_client_row(job, payload)
            elif job.entity_type == ImportJob.EntityTypes.LEADS:
                result = import_lead_row(job, payload)
            elif job.entity_type == ImportJob.EntityTypes.SALES:
                result = import_sale_row(job, payload, connector=connector)
            elif job.entity_type == ImportJob.EntityTypes.CATALOG:
                result = import_catalog_row(job, payload, connector=connector)
            else:
                raise ValidationError("Unsupported import entity.")
            summary[result] = summary.get(result, 0) + 1
    except Exception as exc:
        sync_run.status = ConnectorSyncRun.Statuses.FAILED
        sync_run.error = str(exc)
        sync_run.finished_at = timezone.now()
        sync_run.save(update_fields=["status", "error", "finished_at"])
        connector.status = BusinessConnector.Statuses.FAILED
        connector.last_error = str(exc)
        connector.save(update_fields=["status", "last_error", "updated_at"])
        raise
    imported = summary["created"] + summary["updated"]
    job.status = ImportJob.Statuses.IMPORTED
    job.imported_count = imported
    job.imported_at = timezone.now()
    job.error = ""
    job.preview_json = {
        **(job.preview_json or {}),
        "import_summary": {
            **((job.preview_json or {}).get("import_summary") or {}),
            **summary,
            "total_rows": len(rows),
            "imported": imported,
            "ready": True,
        },
    }
    job.save(update_fields=["status", "imported_count", "imported_at", "error", "preview_json", "updated_at"])
    sync_run.status = ConnectorSyncRun.Statuses.SUCCEEDED
    sync_run.events_received = len(rows)
    sync_run.events_processed = imported
    sync_run.finished_at = timezone.now()
    sync_run.save(update_fields=["status", "events_received", "events_processed", "finished_at"])
    connector.status = BusinessConnector.Statuses.CONNECTED
    connector.last_error = ""
    connector.last_sync_at = sync_run.finished_at
    connector.connected_at = connector.connected_at or timezone.now()
    connector.config_json = {
        **(connector.config_json or {}),
        "last_entity_type": job.entity_type,
        "last_import_job_id": job.id,
        "last_filename": job.original_filename,
        "last_summary": summary,
        "supported_entities": [
            ImportJob.EntityTypes.CLIENTS,
            ImportJob.EntityTypes.LEADS,
            ImportJob.EntityTypes.SALES,
            ImportJob.EntityTypes.CATALOG,
        ],
    }
    connector.save(update_fields=["status", "last_error", "last_sync_at", "connected_at", "config_json", "updated_at"])
    write_audit_log(
        request,
        AuditLog.Actions.CREATE,
        job,
        business=job.business,
        metadata={"kind": "import_confirmed", "entity_type": job.entity_type, "rows": imported},
    )
    return job


def ensure_excel_csv_connector(job):
    connector, _ = BusinessConnector.objects.get_or_create(
        business=job.business,
        provider=BusinessConnector.Providers.EXCEL_CSV,
        name="Excel / CSV",
        defaults={
            "capability": BusinessConnector.Capabilities.SALES,
            "auth_type": BusinessConnector.AuthTypes.NONE,
            "status": BusinessConnector.Statuses.CONNECTED,
            "created_by": job.actor,
            "connected_at": timezone.now(),
            "config_json": {
                "source": "file_import",
                "supported_entities": [
                    ImportJob.EntityTypes.CLIENTS,
                    ImportJob.EntityTypes.LEADS,
                    ImportJob.EntityTypes.SALES,
                    ImportJob.EntityTypes.CATALOG,
                ],
            },
        },
    )
    if connector.status not in {BusinessConnector.Statuses.CONNECTED, BusinessConnector.Statuses.SYNCING}:
        connector.status = BusinessConnector.Statuses.CONNECTED
        connector.last_error = ""
        if connector.connected_at is None:
            connector.connected_at = timezone.now()
        connector.save(update_fields=["status", "last_error", "connected_at", "updated_at"])
    return connector


def mark_excel_csv_import_failed(job, error):
    connector = ensure_excel_csv_connector(job)
    message = str(error)
    run = ConnectorSyncRun.objects.create(
        business=job.business,
        connector=connector,
        mode=ConnectorSyncRun.Modes.MANUAL,
        status=ConnectorSyncRun.Statuses.FAILED,
        started_at=timezone.now(),
        finished_at=timezone.now(),
        events_received=job.total_rows,
        events_processed=0,
        error=message,
    )
    connector.status = BusinessConnector.Statuses.FAILED
    connector.last_error = message
    connector.last_sync_at = run.finished_at
    connector.config_json = {
        **(connector.config_json or {}),
        "last_entity_type": job.entity_type,
        "last_import_job_id": job.id,
        "last_filename": job.original_filename,
        "last_summary": {"created": 0, "updated": 0, "skipped": 0, "error": message},
    }
    connector.save(update_fields=["status", "last_error", "last_sync_at", "config_json", "updated_at"])
    return run


def aliases_for_entity(entity_type):
    aliases = FIELD_ALIASES_BY_ENTITY.get(entity_type)
    if aliases is None:
        raise ValidationError("Unsupported import entity.")
    return aliases


def import_client_row(job, payload):
    if not payload.get("full_name") and not payload.get("phone") and not payload.get("email"):
        return "skipped"
    existing = find_import_client(job.business, payload)
    if existing:
        changed = fill_missing_client_fields(existing, payload)
        return "updated" if changed else "skipped"
    Client.objects.create(business=job.business, **client_defaults(payload))
    return "created"


def import_lead_row(job, payload):
    if not payload.get("full_name") and not payload.get("phone") and not payload.get("email") and not payload.get("message"):
        return "skipped"
    client = find_import_client(job.business, payload)
    if client:
        fill_missing_client_fields(client, payload)
    else:
        client = Client.objects.create(business=job.business, **client_defaults(payload, fallback="Imported lead"))
    service = None
    if payload.get("service_name"):
        service = Service.objects.filter(business=job.business, name__iexact=payload["service_name"]).first()
    status = payload.get("status") if payload.get("status") in dict(Lead.Statuses.choices) else Lead.Statuses.NEW
    source = payload.get("source") if payload.get("source") in dict(Lead.Sources.choices) else Lead.Sources.OTHER
    message = payload.get("message") or "Imported lead"
    if Lead.objects.filter(business=job.business, client=client, service=service, source=source, message=message).exists():
        return "skipped"
    Lead.objects.create(
        business=job.business,
        client=client,
        service=service,
        source=source,
        message=message,
        status=status,
    )
    return "created"


def import_sale_row(job, payload, connector=None):
    sale_payload = normalize_sale_payload(payload)
    occurred_at = sale_payload.pop("occurred_at_parsed", None)
    _, created = normalize_business_event(
        business=job.business,
        source=sale_payload.get("source") or "manual_import",
        event_type="sale.recorded",
        external_id=sale_payload.get("external_id", ""),
        payload=sale_payload,
        connector=connector,
        occurred_at=occurred_at,
    )
    return "created" if created else "skipped"


def import_catalog_row(job, payload, connector=None):
    catalog_payload = normalize_catalog_payload(payload)
    service = None
    service_created = False
    if catalog_payload["item_type"] == "service":
        service, service_created = Service.objects.update_or_create(
            business=job.business,
            name=catalog_payload["name"],
            defaults={
                "description": catalog_payload.get("description", ""),
                "duration_minutes": catalog_payload.get("duration_minutes") or 30,
                "price_from": catalog_payload.get("price_from"),
                "is_active": True,
            },
        )
    if service:
        catalog_payload["service_id"] = service.id
    _, event_created = normalize_business_event(
        business=job.business,
        source=catalog_payload.get("source") or "manual_import",
        event_type="catalog.item_imported",
        external_id=catalog_payload.get("sku", ""),
        payload=catalog_payload,
        connector=connector,
    )
    if event_created:
        return "created"
    if service and service_created:
        return "created"
    return "skipped"


def find_import_client(business, payload):
    query = Q()
    if payload.get("phone"):
        query |= Q(phone=payload["phone"])
    if payload.get("email"):
        query |= Q(email__iexact=payload["email"])
    if not query:
        return None
    return Client.objects.filter(business=business).filter(query).first()


def client_defaults(payload, fallback="Imported client"):
    return {
        "full_name": payload.get("full_name") or payload.get("phone") or payload.get("email") or fallback,
        "phone": payload.get("phone", ""),
        "email": payload.get("email", ""),
        "source": normalize_client_source(payload.get("source")),
        "notes": payload.get("notes", ""),
    }


def fill_missing_client_fields(client, payload):
    changed_fields = []
    for field in ["full_name", "phone", "email", "notes"]:
        value = payload.get(field)
        if value and not getattr(client, field):
            setattr(client, field, value)
            changed_fields.append(field)
    source = normalize_client_source(payload.get("source"))
    if source and source != client.source and client.source == Client.Sources.OTHER:
        client.source = source
        changed_fields.append("source")
    if changed_fields:
        client.save(update_fields=[*changed_fields, "updated_at"])
    return bool(changed_fields)


def normalize_client_source(value):
    return value if value in dict(Client.Sources.choices) else Client.Sources.OTHER


def read_tabular_file(path):
    extension = Path(path).suffix.lower()
    if extension == ".csv":
        with open(path, newline="", encoding="utf-8-sig") as file:
            reader = csv.DictReader(file)
            rows = list(reader)
            if not reader.fieldnames:
                raise ValidationError("Import file is empty.")
            return normalize_tabular_rows(rows)
    if extension == ".xlsx":
        if load_workbook is None:
            raise ValidationError("XLSX import requires openpyxl. Install requirements.txt and retry.")
        workbook = load_workbook(path, read_only=True, data_only=True)
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            raise ValidationError("Import file is empty.")
        headers = [str(value or "").strip() for value in rows[0]]
        result = []
        for values in rows[1:]:
            result.append({headers[index]: cell_to_string(value) for index, value in enumerate(values) if index < len(headers)})
        workbook.close()
        return normalize_tabular_rows(result)
    raise ValidationError("Only CSV and XLSX files are supported.")


def normalize_tabular_rows(rows):
    if not rows:
        raise ValidationError("Import file is empty.")
    normalized_rows = []
    for row in rows:
        clean_row = {str(key or "").strip(): cell_to_string(value) for key, value in (row or {}).items() if str(key or "").strip()}
        if any(value for value in clean_row.values()):
            normalized_rows.append(clean_row)
    if not normalized_rows:
        raise ValidationError("Import file has headers but no data rows.")
    if len(normalized_rows) > settings.IMPORT_MAX_ROWS:
        raise ValidationError(f"Import file has too many rows. Maximum is {settings.IMPORT_MAX_ROWS}.")
    return normalized_rows


def cell_to_string(value):
    if value is None:
        return ""
    return str(value).strip()


def guess_mapping(headers, aliases):
    mapping = {}
    normalized = {normalize_header(header): header for header in headers}
    for field, candidates in aliases.items():
        for candidate in candidates:
            header = normalized.get(normalize_header(candidate))
            if header:
                mapping[field] = header
                break
    return mapping


def normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def mapped_payload(row, mapping, entity_type=ImportJob.EntityTypes.CLIENTS):
    allowed_fields = FIELDS_BY_ENTITY.get(entity_type, CLIENT_FIELDS)
    return {field: str(row.get(header, "") or "").strip() for field, header in mapping.items() if field in allowed_fields}


def duplicate_preview(job, rows, mapping):
    result = []
    seen = {}
    for index, row in enumerate(rows, start=1):
        payload = mapped_payload(row, mapping, job.entity_type)
        row_key = stable_import_row_key(payload)
        if row_key in seen:
            result.append({"row": index, "payload": payload, "duplicates": [{"id": "", "full_name": f"Duplicate row {seen[row_key]}", "matched_fields": ["file_row"]}]})
            continue
        seen[row_key] = index
        duplicates = find_duplicate_clients(job.business, phone=payload.get("phone"), email=payload.get("email"))
        if duplicates:
            result.append({"row": index, "payload": payload, "duplicates": duplicate_payload(duplicates, phone=payload.get("phone"), email=payload.get("email"))})
    return result


def stable_import_row_key(payload):
    normalized = json.dumps(payload or {}, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(normalized.encode("utf-8")).hexdigest()


def validate_import_rows(entity_type, rows, mapping):
    errors = []
    required = {
        ImportJob.EntityTypes.CLIENTS: [],
        ImportJob.EntityTypes.LEADS: ["message"],
        ImportJob.EntityTypes.SALES: ["amount"],
        ImportJob.EntityTypes.CATALOG: ["name"],
    }.get(entity_type, [])
    for field in required:
        if field not in mapping:
            errors.append({"row": 0, "field": field, "message": f"Required column for {field} is missing."})
    for index, row in enumerate(rows, start=1):
        payload = mapped_payload(row, mapping, entity_type)
        if entity_type == ImportJob.EntityTypes.SALES:
            if payload.get("amount") and parse_decimal(payload["amount"]) is None:
                errors.append({"row": index, "field": "amount", "message": "Amount must be a number."})
            if payload.get("quantity") and parse_decimal(payload["quantity"]) is None:
                errors.append({"row": index, "field": "quantity", "message": "Quantity must be a number."})
            if payload.get("occurred_at") and parse_datetime(payload["occurred_at"]) is None:
                errors.append({"row": index, "field": "occurred_at", "message": "Use ISO datetime, for example 2026-05-22T10:00:00+05:00."})
        if entity_type == ImportJob.EntityTypes.CATALOG:
            if not payload.get("name"):
                errors.append({"row": index, "field": "name", "message": "Name is required."})
            if payload.get("price_from") and parse_decimal(payload["price_from"]) is None:
                errors.append({"row": index, "field": "price_from", "message": "Price must be a number."})
            if payload.get("duration_minutes") and parse_positive_int(payload["duration_minutes"]) is None:
                errors.append({"row": index, "field": "duration_minutes", "message": "Duration must be a positive integer."})
            if payload.get("stock_quantity") and parse_decimal(payload["stock_quantity"]) is None:
                errors.append({"row": index, "field": "stock_quantity", "message": "Stock quantity must be a number."})
    return errors[:50]


def readable_import_error(row_errors):
    if not row_errors:
        return ""
    first = row_errors[0]
    return f"Import has {len(row_errors)} error(s). First: row {first.get('row')}, {first.get('field')}: {first.get('message')}"


def normalize_sale_payload(payload):
    amount = parse_decimal(payload.get("amount"))
    if amount is None:
        raise ValidationError({"amount": "Amount must be a number."})
    quantity = parse_decimal(payload.get("quantity")) or Decimal("1")
    occurred_at = parse_datetime(payload.get("occurred_at", "")) if payload.get("occurred_at") else timezone.now()
    if occurred_at is None:
        raise ValidationError({"occurred_at": "Use ISO datetime."})
    return {
        "external_id": payload.get("external_id", ""),
        "occurred_at": occurred_at.isoformat(),
        "occurred_at_parsed": occurred_at,
        "client_name": payload.get("client_name", ""),
        "phone": payload.get("phone", ""),
        "item_name": payload.get("item_name", ""),
        "quantity": str(quantity),
        "amount": str(amount),
        "source": payload.get("source") or "manual",
        "notes": payload.get("notes", ""),
    }


def normalize_catalog_payload(payload):
    name = payload.get("name", "").strip()
    if not name:
        raise ValidationError({"name": "Name is required."})
    item_type = normalize_catalog_item_type(payload.get("item_type"))
    return {
        "item_type": item_type,
        "sku": payload.get("sku", ""),
        "name": name,
        "description": payload.get("description", ""),
        "duration_minutes": parse_positive_int(payload.get("duration_minutes")) or 30,
        "price_from": decimal_to_string(parse_decimal(payload.get("price_from"))),
        "stock_quantity": decimal_to_string(parse_decimal(payload.get("stock_quantity"))),
        "source": payload.get("source") or "manual",
    }


def normalize_catalog_item_type(value):
    normalized = str(value or "service").strip().lower()
    if normalized in {"service", "услуга"}:
        return "service"
    if normalized in {"product", "товар", "stock", "inventory"}:
        return "product"
    return "product"


def parse_decimal(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(" ", "").replace(",", "."))
    except (InvalidOperation, ValueError):
        return None


def parse_positive_int(value):
    if value in (None, ""):
        return None
    try:
        parsed = int(Decimal(str(value).replace(",", ".")))
    except (InvalidOperation, ValueError):
        return None
    return parsed if parsed > 0 else None


def decimal_to_string(value):
    return str(value) if value is not None else ""


def create_manual_sale(business, payload, request):
    sale_payload = normalize_sale_payload(payload)
    occurred_at = sale_payload.pop("occurred_at_parsed", None)
    event, _ = normalize_business_event(
        business=business,
        source=sale_payload.get("source") or "manual",
        event_type="sale.recorded",
        external_id=sale_payload.get("external_id", ""),
        payload=sale_payload,
        occurred_at=occurred_at,
    )
    write_audit_log(request, AuditLog.Actions.CREATE, event, business=business, metadata={"kind": "manual_sale_created"})
    return event


def create_manual_catalog_item(business, payload, request):
    catalog_payload = normalize_catalog_payload(payload)
    if catalog_payload["item_type"] == "service":
        service, _ = Service.objects.update_or_create(
            business=business,
            name=catalog_payload["name"],
            defaults={
                "description": catalog_payload.get("description", ""),
                "duration_minutes": catalog_payload.get("duration_minutes") or 30,
                "price_from": catalog_payload.get("price_from"),
                "is_active": True,
            },
        )
        catalog_payload["service_id"] = service.id
    event, _ = normalize_business_event(
        business=business,
        source=catalog_payload.get("source") or "manual",
        event_type="catalog.item_created",
        external_id=catalog_payload.get("sku", ""),
        payload=catalog_payload,
    )
    write_audit_log(request, AuditLog.Actions.CREATE, event, business=business, metadata={"kind": "manual_catalog_item_created"})
    return event


def import_template_response(entity_type):
    template = IMPORT_TEMPLATES.get(entity_type)
    if template is None:
        raise ValidationError("Unsupported import template.")
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(template["headers"])
    writer.writerow(template["row"])
    response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{entity_type}_template.csv"'
    return response


def export_csv_response(queryset, fields, filename):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(fields)
    for instance in queryset:
        writer.writerow([safe_csv_cell(getattr(instance, field, "")) for field in fields])
    response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def export_clients(business):
    return export_csv_response(
        Client.objects.filter(business=business, is_archived=False).order_by("id"),
        ["id", "full_name", "phone", "email", "source", "notes", "created_at"],
        "clients.csv",
    )


def export_leads(business):
    return export_csv_response(
        Lead.objects.filter(business=business, is_archived=False).order_by("id"),
        ["id", "client_id", "service_id", "source", "status", "message", "responsible_user_id", "created_at"],
        "leads.csv",
    )


def export_deals(business):
    return export_csv_response(
        Deal.objects.filter(business=business, is_archived=False).order_by("id"),
        ["id", "client_id", "pipeline_id", "stage_id", "title", "amount", "status", "owner_id", "source", "created_at"],
        "deals.csv",
    )


def export_sales(business):
    return export_business_events(business, "sale.recorded", "sales.csv")


def export_catalog(business):
    return export_business_events(business, "catalog.", "catalog.csv")


def export_business_events(business, event_type_prefix, filename):
    buffer = io.StringIO()
    fields = ["id", "event_type", "source", "external_id", "occurred_at", "payload_json"]
    writer = csv.writer(buffer)
    writer.writerow(fields)
    queryset = BusinessEvent.objects.filter(business=business, event_type__startswith=event_type_prefix).order_by("id")
    for event in queryset:
        writer.writerow([safe_csv_cell(value) for value in [event.id, event.event_type, event.source, event.external_id, event.occurred_at.isoformat(), event.payload_json]])
    response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response
